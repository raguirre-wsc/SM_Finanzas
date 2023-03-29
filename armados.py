# noinspection PyUnresolvedReferences
from openpyxl import Workbook

# noinspection PyUnresolvedReferences
import ctypes

# noinspection PyUnresolvedReferences
from tkinter import *

# noinspection PyUnresolvedReferences
import openpyxl

# noinspection PyUnresolvedReferences
import arrow

# noinspection PyUnresolvedReferences
import numpy as np

# noinspection PyUnresolvedReferences
from auxiliar import auxiliar

# noinspection PyUnresolvedReferences
import pandas

# noinspection PyUnresolvedReferences
import xlrd

# noinspection PyUnresolvedReferences
import xlwings

# noinspection PyUnresolvedReferences
from openpyxl import load_workbook

# noinspection PyUnresolvedReferences
import xlsxwriter

# noinspection PyUnresolvedReferences
import keyboard

# noinspection PyUnresolvedReferences
import time

# noinspection PyUnresolvedReferences
import os

# noinspection PyUnresolvedReferences
import glob

# noinspection PyUnresolvedReferences
from tkinter import *

# noinspection PyUnresolvedReferences
from PIL import ImageTk, Image


class armadora:
    def aperturaConsolidado():
        #GENERADOR ARCHIVO TRANSFER
        directorioo="C:/Users/rodriaguirre/Desktop/consolidado.xlsx"

        extractototal=xlwings.Book(directorioo)
        extractototal.sheets['Hoja2'].select()
        extractototal.api.RefreshAll()

        ctypes.windll.user32.MessageBoxW(0, "Proceso termiando", "Confirmación", 0)

    def armadorArchExtractos():
        #CREA VERSION XLSX DE EXTRACTO
        extractototal=xlwings.Book("C:/Users/rodriaguirre/OneDrive - Swiss Medical S.A/Documents/General/03. Posicion Financiera Diaria/01. Saldia/01. Salud/movim_extra_mibarberis01.xls")
        extractototal.save("C:/Users/rodriaguirre/OneDrive - Swiss Medical S.A/Documents/General/03. Posicion Financiera Diaria/01. Saldia/01. Salud/movim_extra_mibarberis01.xlsx")
        extractototal.close()


        #FILTRADOR DE EXTRACTO
        extracto_dir="C:/Users/rodriaguirre/OneDrive - Swiss Medical S.A/Documents/General/03. Posicion Financiera Diaria/01. Saldia/01. Salud/movim_extra_mibarberis01.xlsx"
        #extracto_dir="C:/Users/rodrigo/Desktop/movim_extra_rodriaguirre0601.xlsx"
        extracto=pandas.read_excel(extracto_dir)
        extraseguros=extracto.loc[extracto["emp_des_de"]=="SMG COMPAÑIA ARGENTINA DE SEGUROS"]
        extralife=extracto.loc[extracto["emp_des_de"]=="SMG LIFE SEGUROS DE VIDA SA"]
        extraart=extracto.loc[extracto["emp_des_de"]=="SWISS MEDICAL ART"]

        # Create a Pandas Excel writer using XlsxWriter as the engine.
        writer = pandas.ExcelWriter(r'C:\Users\rodriaguirre\Desktop\extractos_mult.xlsx', engine='xlsxwriter')

        # Write each dataframe to a different worksheet.
        extraseguros.to_excel(writer, sheet_name='Generales', index=False, header=False)
        extraart.to_excel(writer, sheet_name='ART', index=False, header=False)
        extralife.to_excel(writer, sheet_name='Life', index=False, header=False)

        # Close the Pandas Excel writer and output the Excel file.
        writer.save()

        ctypes.windll.user32.MessageBoxW(0, "Proceso termiando", "Confirmación", 0)

    def armadorArchTransfer():
        #GENERADOR ARCHIVO TRANSFER
        directorioo="C:/Users/rodriaguirre/OneDrive - Swiss Medical S.A/Documents/General/03. Posicion Financiera Diaria/02. Transferencias/Transferencias basico.xlsm"
        directorioa="C:/Users/rodriaguirre/OneDrive - Swiss Medical S.A/Documents/General/03. Posicion Financiera Diaria/02. Transferencias/"+  arrow.now().format('YYYY') + "/" +  arrow.now().format('MM') + ". " + auxiliar.nombreMes(str(arrow.now().format('MM')))+"/Transferencias "+arrow.now().format('DD-MM-YYYY')+".xlsm"

        libro= openpyxl.load_workbook(filename=directorioo, read_only=False, keep_vba=True)

        libro.save(directorioa)

        ctypes.windll.user32.MessageBoxW(0, "Proceso termiando", "Confirmación", 0)

    def armadorPosicion():
        #GENERADOR ARCHIVO POSICION
        directoriopa = r"C:\Users\rodriaguirre\OneDrive - Swiss Medical S.A\Documents\General/03. Posicion Financiera Diaria/01. Saldia/02. Seguros/Saldia basico.xlsx"
        directoriopn = "C:/Users/rodriaguirre/OneDrive - Swiss Medical S.A/Documents/General/03. Posicion Financiera Diaria/01. Saldia/02. Seguros/" + arrow.now().format('YYYY') + "/" + arrow.now().format('MM.YYYY') + "/" + arrow.now().format('YYYY_MM_DD') + " SMG.xlsx"

        librop = openpyxl.load_workbook(filename=directoriopa)

        librop.save(directoriopn)

        xlwgus = xlwings.Book(r"C:\Users\rodriaguirre\Desktop\consolidado.xlsx")
        xlwfci = xlwings.Book(r"C:/Users/rodriaguirre/OneDrive - Swiss Medical S.A/Documents/General/03. Posicion Financiera Diaria/05. Inversiones/FCI.xlsx")

        my_values = xlwfci.sheets['Posiciones'].range('G2:G60').options(ndim=2).value
        xlwfci.sheets['Posiciones'].range('E2:E60').value = my_values

        xlwfci.sheets['Posiciones'].range('K2').value = arrow.now().format('MM/DD/YYYY')

        xlwposi = xlwings.Book(str(directoriopn))

        xlwposi.app.activate(steal_focus=True)

        time.sleep(2)

        xlwposi.sheets['SALDIA'].range('R4:S250').value = None
        "xlwposi.sheets['Mov. Manuales'].range('E2:F250').value = None"
        xlwposi.sheets['Pegar Consolidado'].range('A2:H450').value = None

        my_values2 = xlwgus.sheets['Hoja2'].range('A2:H450').options(ndim=2).value
        xlwposi.sheets['Pegar Consolidado'].range('A2:H450').value = my_values2

        xlwposi.sheets['Pegar Consolidado'].range('B4').value = arrow.now().format('MM/DD/YYYY')

        xlwposi.api.RefreshAll()

        ctypes.windll.user32.MessageBoxW(0, "Proceso termiando", "Confirmación", 0)

    def armadorPagos():
        directorio1 = "C:/Users/rodriaguirre/OneDrive - Swiss Medical S.A/Documents/General/03. Posicion Financiera Diaria/06. Pagos SAP/zcl04 " + arrow.now().format('DD.MM') + ".XLS"
        directorio2 = "C:/Users/rodriaguirre\OneDrive - Swiss Medical S.A/Documents/General/03. Posicion Financiera Diaria/06. Pagos SAP/Control de Pagos SAP 2.0.xlsx"

        xlwpb = xlwings.Book(directorio1)
        xlwpb.sheets['zcl04 '+ arrow.now().format('DD.MM')].range('D:D').api.Delete()
        xlwpb.sheets['zcl04 '+ arrow.now().format('DD.MM')].range('E:F').api.Delete()
        xlwpb.sheets['zcl04 '+ arrow.now().format('DD.MM')].range('F:G').api.Delete()
        xlwpb.sheets['zcl04 '+ arrow.now().format('DD.MM')].range('R:R').api.Delete()

        xlwpb.sheets['zcl04 '+ arrow.now().format('DD.MM')].range('1:7').api.Delete()

        xlwpa = xlwings.Book(directorio2)

        xlwpa.sheets['ZCL04'].range('A2:AA45000').value.clear

        my_values3 = xlwpb.sheets['zcl04 '+ arrow.now().format('DD.MM')].range('C1:AC45000').options(ndim=2).value
        xlwpa.sheets['ZCL04'].range('A2:AA45000').value = my_values3

        xlwpa.api.RefreshAll()

        print("nice")

        if (xlwpb.sheets['zcl04 '+ arrow.now().format('DD.MM')].range('C1').end('down').row == xlwpa.sheets['ZCL04'].range('A1').end('down').row-1):
            ctypes.windll.user32.MessageBoxW(0, "Proceso termiando. Hay coincidencia.", "Confirmación", 0)
        else:
            ctypes.windll.user32.MessageBoxW(0, "Error!!!! Error!!! REVISAR, no hubo coincidencia!!! Error!!!", "Confirmación", 0)


    def armadorExtractoscash():
        # CONTADOR FILAS EXTRACTO DATANET
        xl = xlwings.Book(r"C:\Users\rodriaguirre\Desktop\extractos_mult.xlsx")
        rownum = xl.sheets['Generales'].range('A1').current_region.last_cell.row
        rownum1 = xl.sheets['ART'].range('A1').current_region.last_cell.row
        rownum2 = xl.sheets['Life'].range('A1').current_region.last_cell.row

        # EXTRACTO GENERALES
        dirgrles = r"C:\Users\rodriaguirre\OneDrive - Swiss Medical S.A\Documents\General\04. Cash Flow\02. Division Seguros\02. SMG Seguros\Extracto Seguros para cash.xlsx"
        xlgrles = xlwings.Book(dirgrles)
        rownumgrles = xlgrles.sheets['BD Datanet'].range('A1').current_region.last_cell.row + 1

        # EXTRACTO ART
        dirart = r"C:\Users\rodriaguirre\OneDrive - Swiss Medical S.A\Documents\General\04. Cash Flow\02. Division Seguros\01. Swiss Medical ART\Extracto ART para cash.xlsx"
        xlart = xlwings.Book(dirart)
        rownumart = xlart.sheets['BD Datanet'].range('A1').current_region.last_cell.row + 1

        # EXTRACTO LIFE
        dirlife = r"C:\Users\rodriaguirre\OneDrive - Swiss Medical S.A\Documents\General\04. Cash Flow\02. Division Seguros\03. SMG Life\Extracto Life 2020.xlsx"
        xllife = xlwings.Book(dirlife)
        rownumlife = xllife.sheets['BD Datanet'].range('A1').current_region.last_cell.row + 1

        # COPIO EXTRACTO DATANET
        rango_datanet = xl.sheets['Generales'].range((1, 1), (rownum, 33)).options(ndim=2).value
        rango_datanet1 = xl.sheets['ART'].range((1, 1), (rownum1, 33)).options(ndim=2).value
        rango_datanet2 = xl.sheets['Life'].range((1, 1), (rownum2, 33)).options(ndim=2).value

        # PEGO EXTRACTO DATANET
        xlgrles.sheets['BD Datanet'].range((rownumgrles, 1), (rownum + rownumgrles, 33)).value = rango_datanet
        xlart.sheets['BD Datanet'].range((rownumart, 1), (rownum1 + rownumart, 33)).value = rango_datanet1
        xllife.sheets['BD Datanet'].range((rownumlife, 1), (rownum2 + rownumlife, 33)).value = rango_datanet2

        # ESTIRO FORMULAS
        rango_formulas = xlgrles.sheets['BD Datanet'].range((rownumgrles - 2, 35), (rownumgrles - 2, 40)).formula
        xlgrles.sheets['BD Datanet'].range((rownumgrles - 2, 35),(rownumgrles + rownum - 1, 40)).formula = rango_formulas

        rango_formulas = xlart.sheets['BD Datanet'].range((rownumart - 2, 35), (rownumart - 2, 40)).formula
        xlart.sheets['BD Datanet'].range((rownumart - 2, 35), (rownumart + rownum1 - 1, 40)).formula = rango_formulas

        rango_formulas = xllife.sheets['BD Datanet'].range((rownumlife - 2, 35), (rownumlife - 2, 40)).formula
        xllife.sheets['BD Datanet'].range((rownumlife - 2, 35), (rownumlife + rownum2 - 1, 40)).formula = rango_formulas

        # REFRESHALL
        xlgrles.api.RefreshAll()
        xlart.api.RefreshAll()
        xllife.api.RefreshAll()

        ctypes.windll.user32.MessageBoxW(0, "Proceso termiando", "Confirmación", 0)

    def transferirFlujos():
        xlwartc= xlwings.Book(r"C:\Users\rodrigo\Swiss Medical S.A\Finanzas UC - General\04. Cash Flow\02. Division Seguros\01. Swiss Medical ART\Cash Swiss Medical ART 2020.xlsx")
        xlwart = xlwings.Book(r"C:\Users\rodrigo\Swiss Medical S.A\Finanzas UC - General\04. Cash Flow\02. Division Seguros\01. Swiss Medical ART\Swiss ART Flujo Caja.xlsx")
        xlwgrlc= xlwings.Book(r"C:\Users\rodrigo\Swiss Medical S.A\Finanzas UC - General\04. Cash Flow\02. Division Seguros\02. SMG Seguros\Cash Seguros 2020.xlsx")
        xlwgrl = xlwings.Book(r"C:\Users\rodrigo\Swiss Medical S.A\Finanzas UC - General\04. Cash Flow\02. Division Seguros\02. SMG Seguros\SMG Seguros Flujo.xlsx")
        xlwcash = xlwings.Book(r"C:\Users\rodrigo\Desktop\Cash.xlsx")
        cash_sol_ART = "ART " + str(auxiliar.nombreMes(xlwart.sheets["ART a enviar"].range("A1").value))
        cash_sol_grl = "Generales " + str(auxiliar.nombreMes(xlwart.sheets["ART a enviar"].range("A1").value))

        xlwcash.sheets.add(cash_sol_ART)
        xlwcash.sheets.add(cash_sol_grl)

        xlwcash.sheets[cash_sol_ART].clear()
        xlwcash.sheets[cash_sol_grl].clear()

        time.sleep(3)

        xlwart.sheets[cash_sol_ART].range('A1:Z34').copy(xlwcash.sheets[cash_sol_ART].range('A1:Z34'))
        xlwcash.sheets[cash_sol_ART].range('A1:Z34').value = xlwcash.sheets[cash_sol_ART].range('A1:Z34').value

        xlwgrl.sheets[cash_sol_grl + ' 21'].range('A1:Z42').copy(xlwcash.sheets[cash_sol_grl].range('A1:Z42'))
        xlwcash.sheets[cash_sol_grl].range('A1:Z42').value = xlwcash.sheets[cash_sol_grl].range('A1:Z42').value

        xlwcash.sheets[cash_sol_ART].range("B:B").columns.autofit()
        xlwcash.sheets[cash_sol_grl].range("B:B").columns.autofit()

        e=2

        xlwcash.sheets[cash_sol_ART]['6:6'].api.RowHeight = 38
        xlwcash.sheets[cash_sol_grl]['6:6'].api.RowHeight = 38

        xlwcash.sheets[cash_sol_ART]['2:2'].api.RowHeight = e
        xlwcash.sheets[cash_sol_grl]['2:2'].api.RowHeight = e

        xlwcash.sheets[cash_sol_ART]['5:5'].api.RowHeight = e
        xlwcash.sheets[cash_sol_grl]['5:5'].api.RowHeight = e

        xlwcash.sheets[cash_sol_ART]['7:7'].api.RowHeight = e
        xlwcash.sheets[cash_sol_grl]['7:7'].api.RowHeight = e

        xlwcash.sheets[cash_sol_ART]['A:A'].api.ColumnWidth = 3.5
        xlwcash.sheets[cash_sol_grl]['A:A'].api.ColumnWidth = 3.5

        xlwcash.sheets[cash_sol_ART]['X:X'].api.ColumnWidth = 2.5
        xlwcash.sheets[cash_sol_grl]['X:X'].api.ColumnWidth = 2.5

        matriz=[xlwcash.sheets[cash_sol_ART],xlwcash.sheets[cash_sol_grl]]

        xlwcash.activate(steal_focus=False)

        for i in matriz:
            i.activate()
            wb = xlwings.books.active
            active_window = wb.app.api.ActiveWindow
            active_window.FreezePanes = False
            active_window.SplitColumn = 2
            active_window.SplitRow = 7
            active_window.FreezePanes = True

        ctypes.windll.user32.MessageBoxW(0, "Proceso termiando", "Confirmación", 0)

    def pagos_mails():
        f = []
        for (dirpath, dirnames, filenames) in os.walk(r"C:\Users\rodriaguirre\OneDrive - Swiss Medical S.A\Documents\General\03. Posicion Financiera Diaria\06. Pagos SAP\Mails de pagos"):
            f.extend(filenames)
            break

    def propuestas_paula():
        pandas.options.display.float_format = '{:,.2f}'.format
        pandas.set_option('display.max_columns', 500)
        pandas.set_option('display.width', 1000)

        list_of_files = glob.glob(r"C:\Users\rodriaguirre\OneDrive - Swiss Medical S.A\Documents\General\03. Posicion Financiera Diaria\06. Pagos SAP\Mails de pagos\*")

        def mi_func(i):
            return i[-15:-5]

        list_of_files.sort(key=lambda x: time.mktime(time.strptime(mi_func(x), "%d.%m.%Y")))
        print(list_of_files)

        def slicer(i):
            if type(i).__name__ != "str":
                return i
            else:
                return i[0:6]

        df_general = pandas.DataFrame({"Sociedad": [], "Banco propio": [], "Propuesta": [], "Nº documento de pago": [], "Fecha valor": [],"Importe pagado en ML": [], "Nombre del receptor del pago": [], "Archivo": [], "Id": []})
        for i in range(len(list_of_files) - 4, len(list_of_files), 1):
            df = pandas.read_excel(str(list_of_files[i]), sheet_name="Base")  # lee los excels de puala
            print(str(list_of_files[i]))
            df["Sociedad"] = df["Sociedad"].fillna(0)
            df.loc[df.Acreedor == "RESCATES", "Propuesta"] = "RESCATES"
            df['Sociedad'] = df['Sociedad'].astype(int)
            df["Id"] = df["Sociedad"].map(str) + " " + df["Banco propio"].map(str) + " " + df["Propuesta"].map(str) + " " + df["Fecha valor"].map(str)  # crea columna con id soc+propuesta
            df["Archivo"] = list_of_files[i][-34:-7]  # crea columna con nombre de archivo de donde proviene la info
            df["Propuesta"] = df["Propuesta"].apply(slicer)  # le saca al campo propuesta la fecha
            df = df[["Sociedad", "Banco propio", "Propuesta", "Nº documento de pago", "Fecha valor", "Importe pagado en ML","Nombre del receptor del pago", "Archivo", "Id"]]  # filtra los campos de interes
            df_filtro = df.drop_duplicates(subset="Id")
            for i in range(len(df_filtro.index)):
                if len((df_general.loc[df_general["Id"] == df_filtro.iat[i, 8]]).index) != 0:
                    df_general = df_general.loc[df_general["Id"] != df_filtro.iat[i, 8]]
            df_general = pandas.concat([df_general, df])  # concatena los dataframes en una sola tabla

        df_general = df_general.set_index("Fecha valor")  # pone como indice la feccha de pago

        print(df_general)

        writer = pandas.ExcelWriter(r"C:\Users\rodriaguirre\Desktop\Base de Pagos.xlsx",engine='xlsxwriter') #se define el escritor
        df_general.to_excel(writer, sheet_name="Hoja1")
        df_general = df_general[(df_general["Importe pagado en ML"]<-3500000)&(df_general["Sociedad"]==7400)]
        print(df_general)
        df_general.to_excel(writer, sheet_name="Hoja2")

        writer.save()

        pp = xlwings.Book(r"C:\Users\rodriaguirre\Desktop\Pagos Mails Paula.xlsx")  # apertura de archivos
        bp = xlwings.Book(r"C:\Users\rodriaguirre\Desktop\Base de Pagos.xlsx")

        time.sleep(3)

        my_values = bp.sheets['Hoja1'].range('A1:I10000').options(ndim=2).value  # copio y pego
        pp.sheets['BD'].range('A1:I10000').value = my_values

        my_values = bp.sheets['Hoja2'].range('A1:I10000').options(ndim=2).value  # copio y pego
        pp.sheets['BD Sin'].range('A1:I10000').value = my_values

        pp.save(r"C:\Users\rodriaguirre\Desktop\Pagos Mails Paula.xlsx")  # guardo tablero de pagos

        time.sleep(3)

        bp.close()

        pp.api.RefreshAll()

        ctypes.windll.user32.MessageBoxW(0, "Proceso termiando", "Confirmación", 0)








