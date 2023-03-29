# noinspection PyUnresolvedReferences
from openpyxl import Workbook

# noinspection PyUnresolvedReferences
import openpyxl

# noinspection PyUnresolvedReferences
import arrow

# noinspection PyUnresolvedReferences
import numpy as np

# noinspection PyUnresolvedReferences
import xlwings as xl




class auxiliar:
    @staticmethod
    def nombreMes(mes_num):

        nombremes=""
        if (mes_num=="01"):
            nombremes="Enero"
        elif (mes_num=="02"):
            nombremes="Febrero"
        elif (mes_num=="03"):
            nombremes="Marzo"
        elif (mes_num=="04"):
            nombremes="Abril"
        elif (mes_num=="05"):
            nombremes="Mayo"
        elif (mes_num=="06"):
            nombremes="Junio"
        elif (mes_num=="07"):
            nombremes="Julio"
        elif (mes_num=="08"):
            nombremes="Agosto"
        elif (mes_num=="09"):
            nombremes="Septiembre"
        elif (mes_num=="10"):
            nombremes="Octubre"
        elif (mes_num=="11"):
            nombremes="Noviembre"
        elif (mes_num=="12"):
            nombremes="Diciembre"


        return nombremes




class auxiliar2:
    @staticmethod
    def contadorTransfer(eleccioncolumna):

        # DIRECTORIOS DE TRANSFERS

        datet = "C:/Users/rodriaguirre/OneDrive - Swiss Medical S.A/Documents/General/03. Posicion Financiera Diaria/02. Transferencias/" + arrow.now().format('YYYY') + "/" + arrow.now().format('MM') + ". " + auxiliar.nombreMes(str( arrow.now().format('MM'))) + "/Transferencias " + arrow.now().format('DD-MM-YYYY') + ".xlsm"
        #LIBROS
        libro2 = openpyxl.load_workbook(filename=datet, read_only=False, keep_vba=True)
        #HOJAS
        hojat = libro2.get_sheet_by_name("Transferencias")


        q=0
        contador3=7
        trans=0
        dc=0
        print(datet)
        if(eleccioncolumna==1):
            dc= 4
        elif (eleccioncolumna==2):
            dc = 5
        print(dc)
        while q <= 635:
            while isinstance(hojat.cell(row=contador3, column=dc).value, str) or hojat.cell(row=contador3, column=dc).value == None and q <= 635:
                q += 1
                contador3 += 1
            q += 1
            trans += 1
            contador3 += 1
        trans -= 1

        return trans




class auxiliar3:
    def contadorTransfersalud(dc):
        #DIRECTORIO DE SALDIA SALUD

        dates2 = r"C:\Users\rodriaguirre\OneDrive - Swiss Medical S.A\Documents\General\03. Posicion Financiera Diaria\01. Saldia\01. Salud\Saldia.xlsx"
        libro3 = openpyxl.load_workbook(dates2)
        hoja3 = libro3.get_sheet_by_name("Posicion Financiera")
        q=0
        contadorsalud=6
        transsalud=0
        #1=smg
        string_int=0
        transdsalud=0
        transcsalud=0
        mds=[]
        mcs=[]
        f=0
        a=0
        while q <= 110:
            try:
                string_int = int(hoja3.cell(row=contadorsalud,column=13).value)
                prueba=False
            except:
                prueba=True
            while prueba or hoja3.cell(row=contadorsalud,column=13).value == None and q <= 110:
                q+=1
                contadorsalud += 1
                try:
                    string_int = int(hoja3.cell(row=contadorsalud, column=13).value)
                    prueba = False
                except:
                    prueba = True
                if (q==110):
                    break
            try:
                string_int = int(hoja3.cell(row=contadorsalud,column=13).value)
                if (hoja3.cell(row=contadorsalud, column=13).value < 0):
                    transcsalud += 1
                    mcs.append([])
                    mcs[f].append(int(hoja3.cell(row=contadorsalud, column=13).value)*-1000)
                    mcs[f].append(hoja3.cell(row=contadorsalud, column=37).value)
                    mcs[f].append(hoja3.cell(row=contadorsalud, column=36).value)
                    f+=1
                elif (hoja3.cell(row=contadorsalud, column=13).value > 0):
                    transdsalud += 1
                    mds.append([])
                    mds[a].append(int(hoja3.cell(row=contadorsalud, column=13).value) * 1000)
                    mds[a].append(hoja3.cell(row=contadorsalud, column=37).value)
                    mds[a].append(hoja3.cell(row=contadorsalud, column=36).value)
                    a+=1
            except:
                "ultimo none"
            q+=1
            contadorsalud+=1
            if(q==110):
                break

            print(f"cargando...")


        if(dc==1):
            return transdsalud
        elif(dc==2):
            return transcsalud
        elif(dc==3):
            return mds
        elif (dc == 4):
            return mcs

