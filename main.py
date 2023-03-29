# noinspection PyUnresolvedReferences
from openpyxl import Workbook

# noinspection PyUnresolvedReferences
import openpyxl

# noinspection PyUnresolvedReferences
import arrow

# noinspection PyUnresolvedReferences
import numpy as np

# noinspection PyUnresolvedReferences
from auxiliar import auxiliar

# noinspection PyUnresolvedReferences
from auxiliar import auxiliar2

# noinspection PyUnresolvedReferences
from auxiliar import auxiliar3

# noinspection PyUnresolvedReferences
import xlwings as xl

# noinspection PyUnresolvedReferences
import ctypes

# noinspection PyUnresolvedReferences
import pandas

matriz_acl=[]

class transferencias:
    def armadorTransferencias():
        #DIRECTORIO DE SALDIA SEGUROS
        dates="C:/Users/rodriaguirre/OneDrive - Swiss Medical S.A/Documents/General/03. Posicion Financiera Diaria/01. Saldia/02. Seguros/" + arrow.now().format('YYYY') + "/" +arrow.now().format('MM.YYYY') + "/" + arrow.now().format('YYYY_MM_DD') + " SMG.xlsx"

        #DIRECTORIO DE TRANSFERS
        datet="C:/Users/rodriaguirre/OneDrive - Swiss Medical S.A/Documents/General/03. Posicion Financiera Diaria/02. Transferencias/"+ arrow.now().format('YYYY') + "/" + arrow.now().format('MM') + ". " +  auxiliar.nombreMes(str(arrow.now().format('MM')))+"/Transferencias "+arrow.now().format('DD-MM-YYYY')+".xlsm"

        #DIRECTORIO DE SALDIA SALUD
        dates2=r"C:\Users\rodriaguirre\OneDrive - Swiss Medical S.A\Documents\General\03. Posicion Financiera Diaria\01. Saldia\01. Salud\Saldia.xlsx"

        print(datet)

        #LECTURA DE EXCELS
        libro =openpyxl.load_workbook(dates)
        hoja = libro.get_sheet_by_name("SALDIA")
        libro2= openpyxl.load_workbook(filename=datet, read_only=False, keep_vba=True)
        hojat=libro2.get_sheet_by_name("Transferencias")
        libro3=openpyxl.load_workbook(dates2)
        hoja3=libro3.get_sheet_by_name("Transferencias")

        print("pase!")
        #VARIABLES
        filasc=hoja.cell(row=3, column=19).value
        filasd=hoja.cell(row=3, column=18).value
        filascs=auxiliar3.contadorTransfersalud(2)
        filasds=auxiliar3.contadorTransfersalud(1)
        columnas=3
        contador=4
        contador2=4
        md=[]
        mc=[]
        mds=auxiliar3.contadorTransfersalud(3)
        mcs=auxiliar3.contadorTransfersalud(4)

        #ARMA MATRIZ DE DEBITOS DESDE SALDIA
        for f in range(filasd+filasds):
            md.append([])
            for c in range(columnas):
                md[f].append(None)
        i=0
        while i<filasd:
            while (hoja.cell(row=contador,column=18).value==None):
                contador += 1
            md[i][0] = hoja.cell(row=contador, column=18).value
            md[i][1] = hoja.cell(row=contador, column=24).value
            md[i][2] = hoja.cell(row=contador, column=27).value
            i=i+1
            contador +=1

        f=0
        q=0
        for f in range(filasd,filasds+filasd):
            md[f][0] = mds[q][0]
            md[f][1] = mds[q][1]
            md[f][2] = mds[q][2]
            q+=1

        #ARMA MATRIZ DE CREDITOS DESDE SALDIA
        for d in range(filasc+filascs):
            mc.append([])
            for e in range(columnas):
                mc[d].append(None)
        j=0
        while j<filasc:
            while (hoja.cell(row=contador2,column=19).value==None):
                contador2 += 1
            mc[j][0] = hoja.cell(row=contador2, column=19).value
            mc[j][1] = hoja.cell(row=contador2, column=24).value
            mc[j][2] = hoja.cell(row=contador2, column=27).value
            j=j+1
            contador2 +=1

        m=0
        d=0
        for d in range(filasc,filascs+filasc):
            mc[d][0] = mcs[m][0]
            mc[d][1] = mcs[m][1]
            mc[d][2] = mcs[m][2]
            m+=1

        #MATRIZ DE AJUSTES DEBITOS
        mad=[]
        tdprevias=auxiliar2.contadorTransfer(1)
        q=0
        e=0
        contador3=7

        for q in range(tdprevias):
            mad.append([])
            for e in range(3):
                mad[q].append(q)

        q=0
        while q<tdprevias:
            while isinstance(hojat.cell(row=contador3, column=4).value,str) or hojat.cell(row=contador3, column=4).value==None:
                contador3+=1
            mad[q][0]=hojat.cell(row=contador3, column=4).value
            mad[q][1] = hojat.cell(row=contador3, column=15).value
            mad[q][2] = hojat.cell(row=contador3, column=19).value
            contador3+=1
            q+=1


        #MATRIZ DE AJUSTES CREDITOS
        mac=[]
        tcprevias=auxiliar2.contadorTransfer(2)
        e=0
        q=0
        contador3=7

        for q in range(tcprevias):
            mac.append([])
            for e in range(3):
                mac[q].append(q)

        q=0
        while q<tcprevias:
            while isinstance(hojat.cell(row=contador3, column=5).value,str) or hojat.cell(row=contador3, column=5).value==None:
                contador3+=1
            mac[q][0]=hojat.cell(row=contador3, column=5).value
            mac[q][1] = hojat.cell(row=contador3, column=15).value
            mac[q][2] = hojat.cell(row=contador3, column=19).value
            contador3+=1
            q+=1
        #show matrices originales
        print("MATRICES ORIGINALES")
        print("debitos")
        q=0
        e=0
        for q in range(filasd+filasds):
            print(f"{md[q][0]}:{md[q][1]}:{md[q][2]}")
        print("creditos")
        q=0
        e=0
        for q in range(filasc+filascs):
            print(f"{mc[q][0]}:{mc[q][1]}:{mc[q][2]}")

        print("------------------------------------")

        #PROCESO DE AJUSTE
        print("EMPIEZA PROCESO DE AJUSTE")
        e=0
        q=0
        for q in range(tdprevias):
            montoajuste=mad[q][0]
            for e in range(filasd):
                if (mad[q][1]==md[e][1]):
                    md[e][0]-=montoajuste

        q=0
        e=0
        montoajuste=0
        for q in range(tcprevias):
            montoajuste=mac[q][0]
            for e in range(filasc):
                if (mac[q][1]==mc[e][1]):
                    mc[e][0]-=montoajuste
        print("terminado")


        #show matrices ajustadas
        print("MATRICES FINALES AJUSTADAS")
        print("debitos")
        q=0
        e=0
        for q in range(filasd):
            print(f"{md[q][0]}:{md[q][1]}:{md[q][2]}")
        print("creditos")
        q=0
        e=0
        for q in range(filasc):
            print(f"{mc[q][0]}:{mc[q][1]}:{mc[q][2]}")

        print("------------------------------------")

        #VARIABLES
        acc1=0
        acc2=0
        acc3=0
        #define primer linea donde escribir
        drow = 81
        crow = 81
        krow = crow
        i = 0
        mu = 0
        while i < 10:
            if (hojat.cell(row=krow, column=5).value == None):
                i += 1
                mu += 1
                krow += 1
            else:
                krow += 2
                drow += mu + 2
                crow += mu + 2
                i += 1
                mu = 0
        filasc=hoja.cell(row=3, column=19).value
        filasd=hoja.cell(row=3, column=18).value
        h=0
        k=0
        y=0
        montodebito=0
        l=0
        acum=0
        acc3=0
        acc2=0
        e=0
        q=0

        global matriz_acl

        #PROCESO PRINCIPAL
        print("CICLO PRINCIPAL")
        for l in range(filasc+filascs):
            contador_acl = 0
            acc1 = mc[l][0]
            acc2 = mc[l][2]
            if (l==0):
                acc3 = mc[l][2]
            else:
                acc3 = mc[l-1][2]
            if(acc2!=acc3):
                if (mc[l][2]==7400):
                    drow = 81
                    crow = 81
                    krow=crow
                    i = 0
                    mu=0
                    while i < 10:
                        if (hojat.cell(row=krow, column=5).value == None):
                            i += 1
                            mu+=1
                            krow+=1
                        else:
                            krow+=2
                            drow += mu+2
                            crow +=mu+2
                            i += 1
                            mu=0
                elif(mc[l][2]==3000):
                    drow=118
                    crow=118
                elif (mc[l][2] == 3800):
                    drow = 155
                    crow = 155
                    krow = crow
                    i = 0
                    mu = 0
                    while i < 10:
                        if (hojat.cell(row=krow, column=5).value == None):
                            i += 1
                            mu += 1
                            krow += 1
                        else:
                            krow += 2
                            drow += mu + 2
                            crow += mu + 2
                            i += 1
                elif (mc[l][2] == 3400):
                    drow = 192
                    crow = 192
                elif (mc[l][2] == 3300):
                    drow = 229
                    crow = 229
                    krow=crow
                    i = 0
                    mu=0
                    while i < 10:
                        if (hojat.cell(row=krow, column=5).value == None):
                            i += 1
                            mu+=1
                            krow+=1
                        else:
                            krow+=2
                            drow += mu+2
                            crow +=mu+2
                            i += 1
                            mu=0
                elif (mc[l][2] == 8100):
                    drow = 266
                    crow = 266
                elif (mc[l][2] == 6000):
                    drow = 303
                    crow = 303
                elif (mc[l][2] == 3700):
                    drow = 340
                    crow = 340
                elif (mc[l][2] == 8200):
                    drow = 377
                    crow = 377
                elif (mc[l][2] == 6500):
                    drow = 414
                    crow = 414
                elif (mc[l][2] == 7300):
                    drow = 451
                    crow = 451
                    krow=crow
                    i = 0
                    mu=0
                    while i < 10:
                        if (hojat.cell(row=krow, column=5).value == None):
                            i += 1
                            mu+=1
                            krow+=1
                        else:
                            krow+=2
                            drow += mu+2
                            crow +=mu+2
                            i += 1
                            mu=0
                elif (mc[l][2] == 7200):
                    drow = 488
                    crow = 488
                    krow=crow
                    i = 0
                    mu=0
                    while i < 10:
                        if (hojat.cell(row=krow, column=5).value == None):
                            i += 1
                            mu+=1
                            krow+=1
                        else:
                            krow+=2
                            drow += mu+2
                            crow +=mu+2
                            i += 1
                            mu=0
                elif (mc[l][2] == 7100):
                    drow = 525
                    crow = 525
                elif (mc[l][2] == 7000):
                    drow = 562
                    crow = 562
                    krow=crow
                    i = 0
                    mu=0
                    while i < 10:
                        if (hojat.cell(row=krow, column=5).value == None):
                            i += 1
                            mu+=1
                            krow+=1
                        else:
                            krow+=2
                            drow += mu+2
                            crow +=mu+2
                            i += 1
                            mu=0
                elif (mc[l][2] == 7600):
                    drow = 599
                    crow = 599
                    krow = crow
                    i = 0
                    mu = 0
                    while i < 10:
                        if (hojat.cell(row=krow, column=5).value == None):
                            i += 1
                            mu += 1
                            krow += 1
                        else:
                            krow += 2
                            drow += mu + 2
                            crow += mu + 2
                            i += 1
                elif(mc[l][2]==1000):
                    drow=7
                    crow=7
                    krow = crow
                    i = 0
                    mu = 0
                    while i < 10:
                        if (hojat.cell(row=krow, column=5).value == None):
                            i += 1
                            mu += 1
                            krow += 1
                        else:
                            krow += 2
                            drow += mu + 2
                            crow += mu + 2
                            i += 1
                elif(mc[l][2]==1100):
                    drow=44
                    crow=44
                    krow = crow
                    i = 0
                    mu = 0
                    while i < 10:
                        if (hojat.cell(row=krow, column=5).value == None):
                            i += 1
                            mu += 1
                            krow += 1
                        else:
                            krow += 2
                            drow += mu + 2
                            crow += mu + 2
                            i += 1
            print(f"empieza el cliclo para {acc1}")
            while acc1!=0:
                filasd=hoja.cell(row=3, column=18).value+filasds
                for i in range(filasd):
                    if (mc[l][2] == md[i][2] and acc1 == md[i][0]):
                        print("entro a ciclo igual, encontre uno igual")
                        hojat.cell(row=drow, column=4).value = acc1
                        hojat.cell(row=drow, column=15).value = md[i][1]
                        crow+=1
                        hojat.cell(row=crow, column=5).value = acc1
                        hojat.cell(row=crow, column=15).value = mc[l][1]
                        drow += 3
                        crow += 2
                        aqui=[mc[l][1],md[i][1],md[i][2]]
                        mc[l][0] = 0
                        md[i][0] = 0
                        acc1 = 0
                        contador_acl+=1
                        break
                    elif (i == filasd-1):
                        print("no encontre uno igual paso a la siguiente fase")
                        break

                filasd=hoja.cell(row=3, column=18).value+filasds
                for k in range(filasd):
                    if (mc[l][2] == md[k][2] and acc1 < md[k][0] and mc[l][0]!=0):
                        print("entre a ciclo mayor, halle un mayor")
                        hojat.cell(row=drow, column=4).value = acc1
                        hojat.cell(row=drow, column=15).value = md[k][1]
                        crow += 1
                        aqui = [mc[l][1],md[k][1],md[k][2]]
                        md[k][0] = md[k][0] - acc1
                        acum+=acc1
                        hojat.cell(row=crow, column=5).value = acum
                        hojat.cell(row=crow, column=15).value = mc[l][1]
                        drow += 3
                        crow += 2
                        mc[l][0] = 0
                        acc1 = 0
                        acum=0
                        contador_acl += 1
                        print(f"el elemento del array ahora vale {mc[l][0]} y {md[k][0]}")
                        if (acc1 == 0):
                            break
                print("pase el ciclo mayor")
                filasd =hoja.cell(row=3, column=18).value+filasds

                toogle=True
                while toogle:
                    for p in range(filasd):
                        if (mc[l][2] == md[p][2] and acc1 > md[p][0] and md[p][0]!=0):
                            print("entre a ciclo menor")
                            hojat.cell(row=drow, column=4).value = md[p][0]
                            hojat.cell(row=drow, column=15).value = md[p][1]
                            drow += 1
                            crow += 1
                            acum+=md[p][0]
                            mc[l][0] = mc[l][0]- md[p][0]
                            acc1 = acc1 - md[p][0]
                            md[p][0] = 0
                            print(acc1)
                            contador_acl += 1
                            if(acc1<=md[p][0]):
                                print("termino fase de menores, paso a buscar el resto igual")
                                toogle=False
                        if (p==filasd-1):
                            toogle=False
                print("pase el ciclo menor")
                filasd =hoja.cell(row=3, column=18).value+filasds
                for i in range(filasd):
                    if (mc[l][2] == md[i][2] and acc1 == md[i][0] and mc[l][0]!=0):
                        print("ok, encontre el resto igual")
                        hojat.cell(row=drow, column=4).value = acc1
                        hojat.cell(row=drow, column=15).value = md[i][1]
                        crow+=1
                        acum+= md[i][0]
                        hojat.cell(row=crow, column=5).value = acum
                        hojat.cell(row=crow, column=15).value = mc[l][1]
                        drow += 3
                        crow += 2
                        aqui = [mc[l][1], md[i][1], md[i][2]]
                        mc[l][0] = 0
                        acc1 = 0
                        md[i][0] = 0
                        acum=0
                        contador_acl += 1
                    elif (i == filasd-1):
                        print("ultima fase no halle resto igual")
                        break
                print("temrine el while")
                print(acc1)
            print(f"el contador aclaraciones es:{contador_acl}")
            if (contador_acl==1):
                matriz_acl.append(aqui)




        print("sali del loop")


        print("-------------y ahora valen-----------")

        print("---------------matriz para aclaraciones------------------")
        print (matriz_acl)
        #for ñ in range(filasc):
         #   print(f"{mc[ñ][0]} - {mc[ñ][2]} / {md[ñ][0]} - {md[ñ][2]}")
        #ñ+=1
        #for ñ in range(ñ,filasd):
        #    print(f"{md[ñ][0]} - {md[ñ][2]}")

        libro2.save("C:/Users/rodriaguirre/OneDrive - Swiss Medical S.A/Documents/General/03. Posicion Financiera Diaria/02. Transferencias/" + arrow.now().format('YYYY') + "/" + arrow.now().format('MM') + ". " + auxiliar.nombreMes(str(arrow.now().format('MM'))) + "/Transferencias " + arrow.now().format('DD-MM-YYYY') + ".xlsm")

        #libro2.save("C:/Users/rodrigo/Swiss Medical S.A/Finanzas UC - General/03. Posicion Financiera Diaria/02. Transferencias/"+auxiliar.nombreMes(str(arrow.now().format('MM')))+"/prueba no usar "+arrow.now().format('DD-MM-YYYY')+".xlsm")

        ctypes.windll.user32.MessageBoxW(0, "Proceso termiando", "Confirmación", 0)


    def transfer_Acl():
        global matriz_acl
#       matriz_acl=[cod_transfer_credito,cod_transfer_debito,cod_soc]

        dir_acc = r"C:\Users\rodrigo\Swiss Medical S.A\Finanzas UC - General\03. Posicion Financiera Diaria\02. Transferencias\Acl_Acc.xlsx"
        dfacc = pandas.read_excel(dir_acc, sheet_name="Hoja2") #abre transferencias accesorias con pandas

        dir="C:/Users/rodrigo/Swiss Medical S.A/Finanzas UC - General/03. Posicion Financiera Diaria/02. Transferencias/Transferencias basico.xlsm"
        df=pandas.read_excel(dir, sheet_name="Base", usecols=[15, 16]) #abre transferencias basico con pandas
        df["Numero"] = df["Numero"].fillna(0)
        df["Numero"] = df["Numero"].astype('int64')

        for i in range(len(matriz_acl)): #para cada elemento de la matriz_acl se fija el cod transfer credito y debito en transfer basico y agrega las cuentas a matriz_acl
                dfs=df.loc[df["Numero"] == matriz_acl[i][0]]
                matriz_acl[i].append("Credito: "+
                dfs.iloc[0][1]
                )
                dfs = df.loc[df["Numero"] == matriz_acl[i][1]]
                matriz_acl[i].append(
                dfs.iloc[0][1]
                )

        dir_acc = r"C:\Users\rodrigo\Swiss Medical S.A\Finanzas UC - General\03. Posicion Financiera Diaria\02. Transferencias\Acl_Acc.xlsx"
        xlacc = xl.Book(dir_acc) #abre transfer accesoria con xlwings

        for i in range(len(matriz_acl)):
            if len(dfacc.loc[dfacc["Codigo"] == matriz_acl[i][0], 'Codigo']) != 0: #si el elemento de matriz_acl se imprimio en DF de transferencias accesorias (osea que es unica cuenta destino)
                xlacc.sheets['Hoja1'].range(i+1,1).value = matriz_acl[i][2] #imprime soc en transferencias accesorias
                xlacc.sheets['Hoja1'].range(i+1,2).value = matriz_acl[i][4] #imprime cuenta debito

        xlacc.save()
        xlacc.close()

        print (matriz_acl)
        print(df)

        ctypes.windll.user32.MessageBoxW(0, "Proceso termiando", "Confirmación", 0)






